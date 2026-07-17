"""
Export GenSoft Desktop (Lamrin SQL Server) masters to GenSoft Order Excel upload formats.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

LogFn = Callable[[str], None]


def _log(log: LogFn | None, msg: str) -> None:
    if log:
        log(msg)


def _cell(val: Any) -> Any:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, bytes):
        return val.decode("utf-8", errors="ignore")
    return val


def _save(path: Path, title: str, headers: list[str], rows: list[list]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for row in rows:
        ws.append([_cell(c) for c in row])
    wb.save(path)
    return path


def connect_sql(cfg: dict) -> Any:
    """Open a pyodbc connection using config keys."""
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is required. Install ODBC Driver 17/18 for SQL Server, then: pip install pyodbc"
        ) from exc

    server = cfg.get("server", ".").strip() or "."
    database = cfg.get("database", "Lamrin").strip() or "Lamrin"
    user = cfg.get("user", "").strip()
    password = cfg.get("password", "")
    trust = str(cfg.get("trust_server_certificate", "true")).lower() in (
        "1",
        "true",
        "yes",
    )
    driver = cfg.get("driver", "ODBC Driver 17 for SQL Server").strip()

    available = [d for d in pyodbc.drivers() if "SQL Server" in d]
    if driver not in available and available:
        driver = available[-1]

    parts = [
        f"DRIVER={{{driver}}}",
        f"SERVER={server}",
        f"DATABASE={database}",
    ]
    if user:
        parts.append(f"UID={user}")
        parts.append(f"PWD={password}")
    else:
        parts.append("Trusted_Connection=yes")
    if trust:
        parts.append("TrustServerCertificate=yes")

    return pyodbc.connect(";".join(parts), timeout=30)


def _fetchall(conn, sql: str, params: tuple = ()) -> list[dict]:
    cur = conn.cursor()
    cur.execute(sql, params)
    cols = [c[0] for c in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


CUSTOMER_HEADERS = [
    "code",
    "name",
    "party_type",
    "address",
    "area",
    "city",
    "mobile",
    "dl_no",
    "gst_no",
    "sales_rep_name",
    "pricing_model",
]

PRODUCT_HEADERS = [
    "product_code",
    "name",
    "manufacturer",
    "pack_size",
    "hsn_code",
    "category",
    "mrp",
    "ptr_rate",
    "pts_rate",
    "gst_pct",
    "batch_no",
    "expiry_date",
    "available_qty",
    "scheme",
    "batch_mrp",
    "batch_ptr_rate",
]

OUTSTANDING_HEADERS = [
    "party_id",
    "party_name",
    "invoice_no",
    "invoice_date",
    "amount",
    "paid",
    "balance",
    "age",
    "discount",
]


def export_customers(conn, out_path: Path, log: LogFn | None = None) -> Path:
    sql = """
    SELECT
        ISNULL(p.partyCode, CAST(p.id AS NVARCHAR(50))) AS code,
        p.partyName AS name,
        CASE
            WHEN UPPER(ISNULL(p.partyType, '')) = 'SUPPLIER' THEN 'supplier'
            ELSE 'customer'
        END AS party_type,
        LTRIM(RTRIM(CONCAT(
            ISNULL(p.address1, ''), ' ',
            ISNULL(p.address2, ''), ' ',
            ISNULL(p.address3, ''), ' ',
            ISNULL(p.address4, '')
        ))) AS address,
        ISNULL(p.place, '') AS area,
        ISNULL(NULLIF(p.city, ''), ISNULL(p.headCity, 'Hyderabad')) AS city,
        ISNULL(NULLIF(p.mobile, ''), ISNULL(p.phone, '')) AS mobile,
        ISNULL(p.dlNo, '') AS dl_no,
        ISNULL(p.gstNo, '') AS gst_no,
        '' AS sales_rep_name,
        'PTR' AS pricing_model
    FROM PartyMaster p
    WHERE ISNULL(p.isActive, 1) = 1
      AND UPPER(ISNULL(p.partyType, '')) NOT IN ('LEDGER')
      AND UPPER(ISNULL(p.category, '')) <> 'LEDGER ACCOUNT'
      AND UPPER(ISNULL(p.partyType, 'BOATH')) IN ('CUSTOMER', 'SUPPLIER', 'BOATH', '')
      AND ISNULL(LTRIM(RTRIM(p.partyName)), '') <> ''
    ORDER BY p.partyName
    """
    rows_raw = _fetchall(conn, sql)
    rows = [
        [
            r["code"],
            r["name"],
            r["party_type"],
            (r["address"] or "").strip(),
            r["area"],
            r["city"] or "Hyderabad",
            r["mobile"],
            r["dl_no"],
            r["gst_no"],
            r["sales_rep_name"],
            r["pricing_model"],
        ]
        for r in rows_raw
    ]
    _log(log, f"Customers: {len(rows)} rows")
    return _save(out_path, "Customers", CUSTOMER_HEADERS, rows)


def export_products(conn, out_path: Path, log: LogFn | None = None) -> Path:
    sql = """
    SELECT
        ISNULL(NULLIF(pm.itemCode, ''), CAST(pm.id AS NVARCHAR(50))) AS product_code,
        pm.itemName AS name,
        ISNULL(m.name, ISNULL(pm.brand, '')) AS manufacturer,
        ISNULL(NULLIF(s.packing, ''), ISNULL(pm.packingUnit, '')) AS pack_size,
        ISNULL(h.code, '') AS hsn_code,
        ISNULL(pm.division, 'General') AS category,
        COALESCE(pm.mrpRate, s.mrp, 0) AS mrp,
        COALESCE(pm.ptrRate, s.ptr, s.retailerPrice, 0) AS ptr_rate,
        COALESCE(pm.ptsRate, s.pts, s.stockistPrice, 0) AS pts_rate,
        COALESCE(s.gstPercent, t.tax, 12) AS gst_pct,
        ISNULL(s.batchNo, '') AS batch_no,
        s.expiryDate AS expiry_date,
        CAST(ROUND(COALESCE(s.aqoh, s.quantityOnHand, 0), 0) AS INT) AS available_qty,
        ISNULL(s.scheme1, '') AS scheme,
        COALESCE(s.mrp, pm.mrpRate, 0) AS batch_mrp,
        COALESCE(s.ptr, s.retailerPrice, pm.ptrRate, 0) AS batch_ptr_rate
    FROM ProductMaster pm
    LEFT JOIN ManufacturerMaster m ON m.id = pm.manufacturerMasterId
    LEFT JOIN HsnMaster h ON h.id = pm.hsnMasterId
    LEFT JOIN TaxMaster t ON t.id = pm.taxMasterId
    LEFT JOIN StockDetail s ON s.productMasterId = pm.id
    WHERE ISNULL(LTRIM(RTRIM(pm.itemName)), '') <> ''
    ORDER BY pm.itemName, s.batchNo
    """
    rows_raw = _fetchall(conn, sql)
    rows = [
        [
            r["product_code"],
            r["name"],
            r["manufacturer"],
            r["pack_size"],
            r["hsn_code"],
            r["category"] or "General",
            float(r["mrp"] or 0),
            float(r["ptr_rate"] or 0),
            float(r["pts_rate"] or 0),
            float(r["gst_pct"] or 12),
            r["batch_no"] or "",
            r["expiry_date"],
            int(r["available_qty"] or 0),
            r["scheme"] or "",
            float(r["batch_mrp"] or 0) or None,
            float(r["batch_ptr_rate"] or 0) or None,
        ]
        for r in rows_raw
    ]
    _log(log, f"Products/stock: {len(rows)} rows")
    return _save(out_path, "Products", PRODUCT_HEADERS, rows)


def export_outstanding_sale_invoices(
    conn, out_path: Path, log: LogFn | None = None
) -> Path:
    """Sale invoices as open bills (paid=0). Best available until invoice aging exists."""
    sql = """
    SELECT
        ISNULL(p.partyCode, CAST(p.id AS NVARCHAR(50))) AS party_id,
        p.partyName AS party_name,
        ISNULL(NULLIF(th.tranNumber, ''), CONCAT('SALE-', th.id)) AS invoice_no,
        CAST(th.tranDate AS DATE) AS invoice_date,
        CAST(ISNULL(th.netAmount, ISNULL(th.grossAmount, 0)) AS FLOAT) AS amount,
        0 AS paid,
        CAST(ISNULL(th.netAmount, ISNULL(th.grossAmount, 0)) AS FLOAT) AS balance,
        NULL AS age,
        CAST(ISNULL(th.discountAmount, 0) AS FLOAT) AS discount
    FROM TranHead th
    INNER JOIN PartyMaster p ON p.id = th.partyMasterId
    WHERE th.tranType IN ('Sale', 'SalesInvoice', 'SI')
      AND ISNULL(th.netAmount, ISNULL(th.grossAmount, 0)) > 0
    ORDER BY th.tranDate DESC, th.id DESC
    """
    rows_raw = _fetchall(conn, sql)
    rows = [
        [
            r["party_id"],
            r["party_name"],
            r["invoice_no"],
            r["invoice_date"],
            float(r["amount"] or 0),
            float(r["paid"] or 0),
            float(r["balance"] or 0),
            "",
            float(r["discount"] or 0),
        ]
        for r in rows_raw
    ]
    _log(log, f"Outstanding (sale invoices): {len(rows)} rows")
    return _save(out_path, "Outstanding", OUTSTANDING_HEADERS, rows)


def export_outstanding_party_balance(
    conn, out_path: Path, log: LogFn | None = None
) -> Path:
    """One synthetic bill per customer using PartyMaster.closingBalance."""
    sql = """
    SELECT
        ISNULL(p.partyCode, CAST(p.id AS NVARCHAR(50))) AS party_id,
        p.partyName AS party_name,
        CONCAT('BAL-', ISNULL(p.partyCode, CAST(p.id AS NVARCHAR(50)))) AS invoice_no,
        CAST(GETDATE() AS DATE) AS invoice_date,
        CAST(ABS(ISNULL(p.closingBalance, 0)) AS FLOAT) AS amount,
        0 AS paid,
        CAST(ABS(ISNULL(p.closingBalance, 0)) AS FLOAT) AS balance,
        NULL AS age,
        0 AS discount
    FROM PartyMaster p
    WHERE ISNULL(p.isActive, 1) = 1
      AND UPPER(ISNULL(p.partyType, '')) NOT IN ('LEDGER')
      AND UPPER(ISNULL(p.category, '')) <> 'LEDGER ACCOUNT'
      AND ABS(ISNULL(p.closingBalance, 0)) > 0.009
      AND (
            UPPER(ISNULL(p.partyType, 'BOATH')) IN ('CUSTOMER', 'BOATH', '')
            OR p.partyType IS NULL
          )
    ORDER BY p.partyName
    """
    rows_raw = _fetchall(conn, sql)
    rows = [
        [
            r["party_id"],
            r["party_name"],
            r["invoice_no"],
            r["invoice_date"],
            float(r["amount"] or 0),
            0,
            float(r["balance"] or 0),
            "",
            0,
        ]
        for r in rows_raw
    ]
    _log(log, f"Outstanding (party balances): {len(rows)} rows")
    return _save(out_path, "Outstanding", OUTSTANDING_HEADERS, rows)


def _xlsx_data_rows(path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        return max((ws.max_row or 1) - 1, 0)
    finally:
        wb.close()


def export_all(
    cfg: dict,
    export_dir: Path,
    log: LogFn | None = None,
) -> dict[str, Path]:
    """Create customers / products / outstanding Excel files."""
    conn = connect_sql(cfg)
    try:
        results: dict[str, Path] = {}
        sync = {
            k.strip().lower()
            for k in (cfg.get("sync_types") or "customers,products,outstanding").split(
                ","
            )
            if k.strip()
        }

        if "customers" in sync:
            results["customers"] = export_customers(
                conn, export_dir / "customers.xlsx", log
            )
        if "products" in sync:
            results["products"] = export_products(
                conn, export_dir / "products_stock.xlsx", log
            )
        if "outstanding" in sync:
            mode = (cfg.get("outstanding_mode") or "sale_invoices").strip().lower()
            out = export_dir / "outstanding.xlsx"
            if mode == "party_balance":
                results["outstanding"] = export_outstanding_party_balance(
                    conn, out, log
                )
            else:
                results["outstanding"] = export_outstanding_sale_invoices(
                    conn, out, log
                )
                if _xlsx_data_rows(out) == 0:
                    _log(log, "No sale invoices found — using party balances")
                    results["outstanding"] = export_outstanding_party_balance(
                        conn, out, log
                    )
        return results
    finally:
        conn.close()
