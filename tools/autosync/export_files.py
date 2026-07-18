"""Discover / convert export files for GenSoft Order upload."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook

LogFn = Callable[[str], None]

# upload_type -> preferred filenames (first match wins)
FILE_CANDIDATES = {
    "customers": [
        "customers.xlsx",
        "customers.xls",
        "customers.csv",
        "customers.tsv",
        "customers.txt",
        "parties.xlsx",
        "parties.xls",
        "parties.csv",
        "parties.txt",
    ],
    "products": [
        "products_stock.xlsx",
        "products_stock.xls",
        "products_stock.csv",
        "products_stock.tsv",
        "products_stock.txt",
        "products.xlsx",
        "products.xls",
        "products.csv",
        "products.tsv",
        "products.txt",
        "stock.xlsx",
        "stock.xls",
        "stock.csv",
        "stock.txt",
    ],
    "outstanding": [
        "outstanding.xlsx",
        "outstanding.xls",
        "outstanding.csv",
        "outstanding.tsv",
        "outstanding.txt",
        "bills.xlsx",
        "bills.xls",
        "bills.csv",
        "bills.txt",
    ],
}


def _log(log: LogFn | None, msg: str) -> None:
    if log:
        log(msg)


def _sniff_delimiter(sample: str) -> str:
    if "\t" in sample:
        return "\t"
    if sample.count(";") > sample.count(","):
        return ";"
    if sample.count("|") > sample.count(","):
        return "|"
    return ","


def delimited_to_xlsx(src: Path, dest: Path) -> Path:
    text = src.read_text(encoding="utf-8-sig", errors="replace")
    if not text.strip():
        raise RuntimeError(f"Empty file: {src}")
    first = text.splitlines()[0] if text.splitlines() else ""
    delim = _sniff_delimiter(first)
    reader = csv.reader(text.splitlines(), delimiter=delim)
    rows = list(reader)
    if not rows:
        raise RuntimeError(f"No rows in {src}")

    wb = Workbook()
    ws = wb.active
    ws.title = dest.stem[:31] or "Sheet1"
    for row in rows:
        ws.append(row)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def dbf_to_xlsx(src: Path, dest: Path) -> Path:
    """Convert a FoxPro/dBASE DBF table → .xlsx.

    VFP `COPY TO file.xls` WITHOUT `TYPE XL5` writes a DBF with an .xls name,
    so sync must handle this format too.
    """
    import struct
    from datetime import date

    data = src.read_bytes()
    nrec = struct.unpack("<I", data[4:8])[0]
    hdr_len = struct.unpack("<H", data[8:10])[0]
    rec_len = struct.unpack("<H", data[10:12])[0]

    fields = []
    pos = 32
    while pos < hdr_len - 1 and data[pos] != 0x0D:
        fname = data[pos : pos + 11].split(b"\x00")[0].decode("ascii", "replace")
        ftype = chr(data[pos + 11])
        flen = data[pos + 16]
        fdec = data[pos + 17]
        fields.append((fname, ftype, flen, fdec))
        pos += 32
    if not fields:
        raise RuntimeError(f"No DBF fields found in {src.name}")

    wb = Workbook()
    ws = wb.active
    ws.title = dest.stem[:31] or "Sheet1"
    ws.append([f[0].lower() for f in fields])

    for i in range(nrec):
        off = hdr_len + i * rec_len
        rec = data[off : off + rec_len]
        if rec[0:1] == b"*":  # deleted record
            continue
        p = 1
        row = []
        for _fname, ftype, flen, fdec in fields:
            raw = rec[p : p + flen]
            p += flen
            text = raw.decode("cp1252", "replace").strip()
            if ftype == "N" and text:
                try:
                    row.append(float(text) if fdec else int(text))
                except ValueError:
                    row.append(text)
            elif ftype == "D" and len(text) == 8 and text.isdigit():
                try:
                    row.append(date(int(text[:4]), int(text[4:6]), int(text[6:8])))
                except ValueError:
                    row.append(text)
            elif ftype == "L":
                row.append(text.upper() in ("T", "Y"))
            else:
                row.append(text)
        ws.append(row)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def _looks_like_dbf(src: Path) -> bool:
    try:
        head = src.open("rb").read(1)
    except OSError:
        return False
    return bool(head) and head[0] in (0x02, 0x03, 0x30, 0x31, 0x83, 0x8B, 0xF5, 0xFB)


def xls_to_xlsx(src: Path, dest: Path) -> Path:
    """Convert legacy Excel .xls → .xlsx (VFP often writes .xls)."""
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xls requires xlrd. Reinstall GenSoftSync / pip install xlrd."
        ) from exc

    # VFP `COPY TO x.xls` (no TYPE) actually writes a DBF table
    if _looks_like_dbf(src):
        return dbf_to_xlsx(src, dest)

    book = xlrd.open_workbook(str(src))
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    ws.title = (dest.stem[:31] or "Sheet1")

    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
                    if (
                        isinstance(dt, datetime)
                        and dt.hour == 0
                        and dt.minute == 0
                        and dt.second == 0
                        and dt.microsecond == 0
                    ):
                        row.append(dt.date())
                    else:
                        row.append(dt)
                except Exception:
                    row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(int(v) if float(v).is_integer() else v)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append("")
            else:
                row.append(cell.value)
        ws.append(row)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def ensure_xlsx(src: Path, log: LogFn | None = None) -> Path:
    """Return an .xlsx path, converting .xls / .csv / .txt when needed."""
    suffix = src.suffix.lower()
    if suffix == ".xlsx":
        return src
    dest = src.with_suffix(".xlsx")
    if suffix == ".xls":
        _log(log, f"Converting {src.name} → {dest.name}")
        return xls_to_xlsx(src, dest)
    if suffix in (".csv", ".tsv", ".txt"):
        _log(log, f"Converting {src.name} → {dest.name}")
        return delimited_to_xlsx(src, dest)
    raise RuntimeError(f"Unsupported file type: {src.name} (use .xlsx or .xls)")


def resolve_export_files(
    export_dir: Path,
    sync_types: str,
    log: LogFn | None = None,
) -> dict[str, Path]:
    """
    Find customers/products/outstanding files in export_dir.
    Converts .xls / .csv / .tsv / .txt → .xlsx when needed.
    """
    export_dir.mkdir(parents=True, exist_ok=True)
    wanted = {
        k.strip().lower()
        for k in (sync_types or "customers,products,outstanding").split(",")
        if k.strip()
    }
    found: dict[str, Path] = {}

    for utype in ("customers", "products", "outstanding"):
        if utype not in wanted:
            continue
        hit: Path | None = None
        for name in FILE_CANDIDATES[utype]:
            p = export_dir / name
            if p.is_file():
                hit = p
                break
        if not hit:
            _log(log, f"Missing file for {utype} in {export_dir}")
            continue
        xlsx = ensure_xlsx(hit, log=log)
        found[utype] = xlsx
        _log(log, f"Using {xlsx.name} for {utype}")

    return found


def run_external_export(
    command: str,
    working_dir: str | Path,
    timeout_sec: int = 600,
    log: LogFn | None = None,
) -> None:
    """Run VFP-built EXE / BAT / CMD that writes export files."""
    import subprocess

    cmd = (command or "").strip()
    if not cmd:
        raise RuntimeError(
            "No export command configured. "
            "Set [external] command to your VFP EXE or BAT that creates the Excel/CSV files."
        )
    cwd = Path(working_dir) if working_dir else Path.cwd()
    cwd.mkdir(parents=True, exist_ok=True)
    _log(log, f"Running export: {cmd}")
    _log(log, f"Working folder: {cwd}")
    result = subprocess.run(
        cmd,
        shell=True,
        cwd=str(cwd),
        capture_output=True,
        text=True,
        timeout=max(30, int(timeout_sec or 600)),
    )
    if result.stdout:
        _log(log, result.stdout.strip()[-2000:])
    if result.stderr:
        _log(log, result.stderr.strip()[-1000:])
    if result.returncode != 0:
        raise RuntimeError(
            f"Export command failed (exit {result.returncode}). "
            "Check the EXE/PRG path and that it writes files into the export folder."
        )
    _log(log, "Export command finished OK")
