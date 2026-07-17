from io import BytesIO
from typing import List

from openpyxl import Workbook, load_workbook


def normalize_header(h: str) -> str:
    key = (h or "").strip().lower().replace(" ", "_")
    aliases = {
        "accounttype": "account_type",
        "business_name": "name",
        "businessname": "name",
        "customer_name": "name",
        # keep party_name as party_name (needed for outstanding);
        # customers upload also accepts party_name via row.get fallback
        "party_code": "code",
        "customer_code": "code",
        "customer_id": "party_id",
        "invoice_number": "invoice_no",
        "inv_no": "invoice_no",
        "inv_date": "invoice_date",
        "qty": "available_qty",
        "stock": "available_qty",
        "quantity": "available_qty",
        "expiry": "expiry_date",
        "exp_date": "expiry_date",
        "product_name": "name",
        "sku": "product_code",
        "gst": "gst_pct",
        "gst%": "gst_pct",
        "sales_rep": "sales_rep_name",
        "owner": "owner_name",
        "dlno": "dl_no",
        "gstno": "gst_no",
    }
    return aliases.get(key, key)


def _xls_to_xlsx_bytes(content: bytes) -> bytes:
    """Convert legacy .xls bytes to .xlsx for openpyxl."""
    import xlrd
    from datetime import datetime

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
                    if (
                        isinstance(dt, datetime)
                        and dt.hour == 0
                        and dt.minute == 0
                        and dt.second == 0
                    ):
                        row.append(dt.date())
                    else:
                        row.append(dt)
                except Exception:
                    row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                v = cell.value
                row.append(int(v) if float(v).is_integer() else v)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(cell.value)
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel_upload(content: bytes) -> List[dict]:
    # Legacy OLE .xls (VFP / Excel 97-2003)
    if content[:8].startswith(b"\xd0\xcf\x11\xe0"):
        try:
            content = _xls_to_xlsx_bytes(content)
        except Exception as exc:
            raise ValueError(f"Could not read .xls file: {exc}") from exc

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    headers = [normalize_header(str(h or "")) for h in header_row]
    parsed = []
    for row in rows_iter:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for key, val in zip(headers, row):
            if key:
                item[key] = "" if val is None else val
        parsed.append(item)
    wb.close()
    return parsed
