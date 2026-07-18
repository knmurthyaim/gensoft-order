from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .. import crud, models, schemas
from ..crud import _normalize_header
from ..database import get_db
from ..deps import require_platform_admin

router = APIRouter(prefix="/api/admin", tags=["admin"])

TEMPLATE_HEADERS = [
    "account_type",
    "name",
    "owner_name",
    "address",
    "area",
    "city",
    "mobile",
    "dl_no",
    "gst_no",
    "email",
    "username",
    "password",
]

SAMPLE_ROW = [
    "retailer",
    "Sample Medical Store",
    "John Doe",
    "Main Road",
    "Kukatpally",
    "Hyderabad",
    "9876543210",
    "TS/HYD/20R-0001",
    "36AAAAA0000A1Z5",
    "sample@example.com",
    "sampleuser",
    "pass1234",
]


@router.get("/accounts", response_model=List[schemas.AdminAccountRow])
def list_accounts(
    search: Optional[str] = Query(None, description="Filter by code, name, username, mobile…"),
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    return crud.list_admin_accounts(db, search=search)


@router.get("/accounts/upload/template")
def download_template(_admin=Depends(require_platform_admin)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(TEMPLATE_HEADERS)
    ws.append(SAMPLE_ROW)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=gensoft_users_template.xlsx"
        },
    )


@router.post("/accounts/upload", response_model=schemas.BulkUploadResult)
async def upload_accounts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    headers = [_normalize_header(str(h or "")) for h in header_row]
    parsed = []
    for row in rows_iter:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for key, val in zip(headers, row):
            if key:
                item[key] = "" if val is None else val
        parsed.append(item)
    wb.close()
    if not parsed:
        raise HTTPException(status_code=400, detail="No data rows found in Excel file")
    return crud.bulk_upload_accounts(db, parsed)


@router.get("/accounts/{account_id}", response_model=schemas.AdminAccountRow)
def get_account(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    row = crud.admin_get_account(db, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Account not found")
    return row


@router.post("/accounts", response_model=schemas.AdminAccountRow, status_code=201)
def create_account(
    data: schemas.RegisterRequest,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    if db.query(models.User).filter(models.User.username == data.username).first():
        raise HTTPException(status_code=400, detail="Username already taken")
    account, user = crud.register_account(db, data)
    return schemas.AdminAccountRow(
        account=account,
        user_id=user.id,
        username=user.username,
        user_name=user.name,
        user_is_active=user.is_active,
    )


@router.put("/accounts/{account_id}", response_model=schemas.AdminAccountRow)
def update_account(
    account_id: int,
    data: schemas.AdminAccountUpdate,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    row = crud.admin_update_account(db, account_id, data)
    if not row:
        raise HTTPException(status_code=404, detail="Account not found")
    return row


@router.delete("/accounts/{account_id}")
def delete_account(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    ok = crud.admin_delete_account(db, account_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"status": "ok", "message": "Account deleted"}


@router.put("/users/{user_id}", response_model=schemas.AdminAccountRow)
def update_user(
    user_id: int,
    data: schemas.AdminUserUpdate,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    try:
        row = crud.admin_update_user(db, user_id, data)
    except crud.AppError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not row:
        raise HTTPException(status_code=404, detail="User not found")
    return row


@router.get("/accounts/{account_id}/data-summary", response_model=schemas.AdminDataSummary)
def account_data_summary(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    row = crud.admin_account_data_summary(db, account_id)
    if not row:
        raise HTTPException(status_code=404, detail="Account not found")
    return row


@router.get("/accounts/{account_id}/products", response_model=List[schemas.Product])
def list_account_products(
    account_id: int,
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    if not crud.admin_get_account(db, account_id):
        raise HTTPException(status_code=404, detail="Account not found")
    return crud.admin_list_products(db, account_id, search=search)


@router.get("/accounts/{account_id}/parties", response_model=List[schemas.Party])
def list_account_parties(
    account_id: int,
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    if not crud.admin_get_account(db, account_id):
        raise HTTPException(status_code=404, detail="Account not found")
    return crud.admin_list_parties(db, account_id, search=search)


@router.get(
    "/accounts/{account_id}/outstanding",
    response_model=List[schemas.OutstandingBillRow],
)
def list_account_outstanding(
    account_id: int,
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    if not crud.admin_get_account(db, account_id):
        raise HTTPException(status_code=404, detail="Account not found")
    bills = crud.admin_list_outstanding(db, account_id, search=search)
    place_map = crud._party_place_by_code(db, [b.party_id for b in bills])
    return [
        schemas.OutstandingBillRow(
            id=b.id,
            party_id=b.party_id,
            party_name=b.party_name,
            place=place_map.get((b.owner_account_id, (b.party_id or "").strip()), ""),
            invoice_no=b.invoice_no,
            invoice_date=b.invoice_date,
            amount=b.amount,
            paid=b.paid,
            balance=b.balance,
            age=crud._bill_age(b.invoice_date, b.age),
            discount=b.discount,
        )
        for b in bills
    ]


@router.delete(
    "/accounts/{account_id}/products",
    response_model=schemas.AdminClearResult,
)
def clear_account_products(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    try:
        return crud.admin_clear_products(db, account_id)
    except crud.AppError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.delete(
    "/accounts/{account_id}/parties",
    response_model=schemas.AdminClearResult,
)
def clear_account_parties(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    try:
        return crud.admin_clear_parties(db, account_id)
    except crud.AppError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.delete(
    "/accounts/{account_id}/outstanding",
    response_model=schemas.AdminClearResult,
)
def clear_account_outstanding(
    account_id: int,
    db: Session = Depends(get_db),
    _admin=Depends(require_platform_admin),
):
    try:
        return crud.admin_clear_outstanding(db, account_id)
    except crud.AppError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
